# Author:   jingnan
# Date:    2021/6/22
# Desc:

from openpyxl import load_workbook

# 抽象成一个类
class ExcelUtil():

    def __init__(self, file, sheet_name):
        self.file = file
        self.sheet_name = sheet_name
        # 获取一个表单对象
        self.sheet_obj = load_workbook(self.file)[self.sheet_name]
        self.max_row = self.sheet_obj.max_row

    # 调用该方法获得数据后注意数据类型，需要使用eval()函数转换为原来的数据类型
    def get_param(self, i, j):

        return self.sheet_obj.cell(i, j).value





# 测试代码
if __name__ == '__main__':

    # 实例化对象
    bravo = ExcelUtil('bravo.xlsx', 'test')
    # 调用函数
    params = bravo.get_param(1, 1)
    print(params)