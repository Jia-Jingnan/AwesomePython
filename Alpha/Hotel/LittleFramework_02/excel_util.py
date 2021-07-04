# Author:   jingnan
# Date:    2021/6/22
# Desc:

from openpyxl import load_workbook
from Alpha.Hotel.LittleFramework_02.config_util import ConfigUtil

# 抽象成一个类
class ExcelUtil():

    def __init__(self, file, sheet_name):
        self.file = file
        self.sheet_name = sheet_name

    # 获取第一行的描述信息，标题行
    def get_header(self):
        # 存储标题行的列表
        workbook = load_workbook(self.file)

        # 定位sheet页
        sheet = workbook[self.sheet_name]
        header = []
        for i in range(1, sheet.max_column + 1):
            header.append(sheet.cell(1, i).value)
        return header

    # 调用该方法获得数据后注意数据类型，需要使用eval()函数转换为原来的数据类型
    # 用例的可配执行
    def get_param(self, case_ids='all'):
        cases = []
        case_ids = ConfigUtil().get_config('golf.config', 'CASE_IDS', 'case_ids')


        # 配置文件读取要执行id列表
        workbook = load_workbook(self.file)

        # 定位sheet页
        sheet = workbook[self.sheet_name]

        # 错误演示1:只定位到单元格，并未取值
        # res = sheet.cell(1,1)
        # print(res)


        # 存储所有测试用例数据的列表
        params = []

        for i in range(2, sheet.max_row + 1):

            # 定义存储一个测试用例数据的字典
            param = {}

            header = self.get_header()

            for j in range(1, sheet.max_column + 1):
                param[header[j-1]] = sheet.cell(i, j).value

            # 将所有的测试用例添加到测试用例列表中
            params.append(param)
        # 筛选指定要执行的case id
        print(type(case_ids))
        if case_ids == 'all':
            cases = params
        else:  # 如果case_ids为id的列表[1,2,4]
            for case in params:
                if case['id'] in case_ids:
                    cases.append(case)

        return cases



# 测试代码
if __name__ == '__main__':

    # 实例化对象
    bravo = ExcelUtil('bravo.xlsx', 'test')
    # 调用函数
    params = bravo.get_param()
    print(params)

