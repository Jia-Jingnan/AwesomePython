# Author:   jingnan
# Date:    2021/6/29
# Desc:

from openpyxl import load_workbook


workbook = load_workbook('bravo.xlsx')

sheet = workbook['test']

res = sheet.cell(1, 1).value
print(res)