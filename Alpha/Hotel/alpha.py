# Author:   jingnan
# Date:    2021/6/21
# Desc:
from openpyxl import load_workbook


# 打开excel
bravo_workbook = load_workbook('bravo.xlsx')


# 指定操作的表单
sheet = bravo_workbook['test']


# 定位单元格， 根据行 列值定位
value = sheet.cell(1, 1).value

print(value)

# 获取最大行
print(sheet.max_row)
# 最大列
print(sheet.max_column)

# 打印出行列值, 数据从excel中取出是什么类型？
# 数字还是数字，其他类型变为字符串类型
print("url：{}, type:{}".format(sheet.cell(1, 1).value, type(sheet.cell(1, 1).value)))
print("type：{}, type:{}".format(sheet.cell(1, 2).value, type(sheet.cell(1, 2).value)))
print("data：{}, type:{}".format(sheet.cell(1, 3).value, type(sheet.cell(1, 3).value)))
print("excepted：{}, type:{}".format(sheet.cell(1, 4).value, type(sheet.cell(1, 4).value)))

# 类型转换， eval()把数据转换成原本的数据类型
s = 'True'
s = '{"age": 18}'
print(type(s))
print(type(eval(s)))

